from openpyxl import Workbook
# from openpyxl.compat import range
from openpyxl.utils import get_column_letter
import preparedata

class WriteExcel(object):

    __col_table_start = 5

    def __init__(self):
        pass

    # エクセルを出力する
    def write_excel(self, data:list, tables, fullname = r'./output.xlsx'):

        sheet_index = 0
        header_row_index = 3
        data_row_index = 4
        col_path = 3
        col_filename = 4
        col_table_index = self.__col_table_start

        workbook = Workbook()
        worksheet = workbook.create_sheet('sheet{0}'.format(sheet_index  + 1))

        # ヘッダーを出力
        worksheet.cell(row = header_row_index, column = col_path).value = 'フォルダ'
        worksheet.cell(row = header_row_index, column = col_filename).value = 'ファイル'
        #       Table名
        idx_tbl_header = self.__col_table_start
        for tablename in tables:
            worksheet.cell(row = header_row_index, column = idx_tbl_header).value = tablename
            idx_tbl_header+=1

        # ボディーを出力
        for result_row in data:

            worksheet.cell(row = data_row_index, column = col_path).value = result_row.path
            worksheet.cell(row = data_row_index, column = col_filename).value = result_row.filename

            col_table_index = self.__col_table_start
            # for useflag in result_row.use_flag_dic.values():
            for tablename in tables:
                useflag = result_row.use_flag_dic[tablename]
                worksheet.cell(row = data_row_index, column = col_table_index).value = '〇' if useflag else ''
                col_table_index += 1

            data_row_index += 1
        try:
            workbook.save('./table_usage.xlsx')
        except IOError:
            print("エクセル「{0}」の出力で失敗しました。\n".format(fullname))
        else:
            print("エクセル「{0}」を出力完了しました。".format(fullname))
        finally:
            workbook.close()
